from openpyxl import load_workbook
from  eTraveler.clientAPI.connection import Connection
import yaml, os, sys


if len(sys.argv) > 1:
    savedir = sys.argv[1]
else:
    savedir = 'test'

excel_file='AutoChanger_hyper_simple.xlsx'
subsystem = 'CHGR'
if not os.path.exists(savedir):
    os.mkdir(savedir)
print "yaml files will be saved in dir ", savedir

myConn = Connection('jrb', 'Raw', prodServer=True)


class yamlAssembly():
    """Main class of this module.

    Provide tools to create yaml files to assembly Filter Exchange System hardware.
    Yaml files are created from a template and excel file provided by meca team.
    """
    def __init__(self, record):
        self._yaml = yaml.load(open("assembly_template.yml"))
        self._name = (record[3].value).encode('utf-8')
        self._yaml['Name'] = subsystem + "_Assembly_" + self._name
        self._yaml['Subsystem'] = subsystem
        self._yaml['HardwareGroup'] = self._name
        self._yaml['Description'] = "Process to assembly " + self._name
        self._yaml['ShortDescription'] = "Assembly of a Exchange System hardware " + self._name
        self.registerHardwareType(self._name, record)

    def registerHardwareType(self, name, rec):
        """Register a HardwareType in database.

        If HardwareType exits with same name in database, just print a message.
        """
        newId = ''
        try: 
            newId = myConn.defineHardwareType(name=name, 
                                              description=rec[7].value,
                                              subsystem=subsystem,
                                              batchedFlag=0,
                                              sequenceWidth='0')
            print 'New hardware type defined.  Returned id is ', newId
            print 'Name=', name
            print 'Description=', description
        except Exception,msg:
            if msg.message != "A component type with name %s already exists."%name:
                print 'Hardware type %s definition failed with exception : %s'%(name,msg.message)

    def registerRelationship(self, relname, name):
        """Define and register a Relationship in database.

        If Relationship exits with same name in database, just print a message.
        """
        newId = ''
        try:
            newId = myConn.defineRelationshipType(name = relname, 
                                                  description = 'rel type via eT API by FV',
                                                  hardwareTypeName = self._name,
                                                  numItems = 1,
                                                  minorTypeName = name,
                                                  slotNames = 'na'
                                                  )
            print 'New relationship defined with id ', newId
        except Exception, msg:
            if "A relationship type with name %s"%relname not in msg.message:
                print 'Relationship definition %s failed with exception : %s'%(relname,msg.message)
        

    def save(self, ingest=False):
        """Save eTraveler assembly in yaml file.
        """
        filepath=os.path.join(savedir,'%s.yaml'%self._name)
        filename = '%s.yaml'%self._name
        print "SAVING file:"
        print "filepath=", filepath
        print "filename=", filename
        with open(filepath, 'w') as outfile:
            # outfile.write( os.path.join(savedir,yaml.dump(self._yaml, default_flow_style=False)) )
            outfile.write(yaml.dump(self._yaml, default_flow_style=False))
            if ingest is True:
                msg = myConn.validateYaml(filepath)
                if msg==0:
                    errorCode = myConn.uploadYaml(filename, reason="")
                    if errorCode !=0:
                        print "ingestion failed with code :", errorCode
                else:
                    "file %s validation failed with error code: %d"%(self._name, msg)
            
    def addRelationship(self, gg):
        """ Add a relationship parent-child.

        Register level 3 hardware type.
        Register Relationship and add a relationship in yaml flow.
        """
        name = (gg[3].value).encode('utf-8')
        if gg[0].value == 3:
            self.registerHardwareType(name, gg)#the deepest elements need registration
        sequence = self._yaml['Sequence'][3]
        print 'sequence=', sequence
        relname = '%s-%s'%(self._name, name)
        new_relationship = {'RelationshipName':relname, 'RelationshipAction':'install'}
        self.registerRelationship(relname, name)
        if sequence['RelationshipTasks'] is None:
            sequence['RelationshipTasks'] = [new_relationship]
        else:
            sequence['RelationshipTasks'].append(new_relationship)


try :
    print "loading the workbook ", excel_file
    wb = load_workbook(filename = excel_file, read_only=True)
    ws = wb.active
    gen = ws.iter_rows()
except:
    wb = load_workbook(filename = excel_file)
    ws = wb.active
    gen = ws.rows

print "excel sheet %s loaded"%excel_file
#yaml_struct = []
#gen=ws.iter_rows()
d0 = None
d1 = None
d2 = None
for gg in gen:
    if gg[3].value == None:
        continue
    elif gg[3].value == "END":
        # it's END of FILE so we have to save the last d2 and d1 yaml flow.
        print "END OF FILE"
        d2.save()
        d1.save()
        break
    print "READING => ",gg[3].value
    if gg[0].value == 0:
        d0 = yamlAssembly(gg)
    if gg[0].value == 1:
        print "### d1 = ", d1
        print "###################"
        if d1 is not None:
            d1.save()
        d0.addRelationship(gg)
        d1 = yamlAssembly(gg)
    if gg[0].value == 2:
        print "### d2 = ", d2
        print "####################"
        if d2 is not None:
            d2.save()
        d1.addRelationship(gg)
        d2 = yamlAssembly(gg)
    if gg[0].value == 3:
        d2.addRelationship(gg)
d0.save(ingest=False)
    #     print gg[3].value 
    #     for gg in gen:
    #         if gg[0].value==2:
    #             print '\t', gg[3].value
    #             for gg in gen:
    #                 if gg[0].value==3 or gg[0].value==None:
    #                     print '\t\t', gg[3].value
    #                 else:
    #                     break
    #         else:
    #             break
